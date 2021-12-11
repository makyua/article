# def main():    
#     import xlwings as xw

#     import requests
#     import json
#     import time
#     #https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db={database}&term={query}
#     #database: 検索するデータベースを指定する
#     #query: 検索するキーワードを指定
#     #retmax: 最大の結果数
#     #retmode: 結果の形式
#     #リファレンス: https://www.ncbi.nlm.nih.gov/books/NBK25499/#chapter4.ESearch
#     #pt: 出版形式
#     tm_st = time.time()
#     query_set_1 = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=pubmed&term="
#     query_set_2 = "&retmax=5&retmode=json"
#     # query_set_2 = "AND review[pt]&retmax=5&retmode=json"
#     import openpyxl
#     #リファレンス: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.reader.excel.html#openpyxl.reader.excel.load_workbook
#     EXCEL_PATH = "C:\\Users\\makyua\\anaconda3\\envs\\gpu\\Scripts\\test\\pubmed.xlsm"
#     #Excelファイルの読み込み
#     wb = xw.Book(EXCEL_PATH)
#     #Excelシートの読み込み
#     ws = wb.sheets["Sheet1"]
#     #キーワードの取得
#     key_1 = ws["B2"].value
#     key_2 = ws["C2"].value
#     key_3 = ws["D2"].value
#     #検索を始める
#     query = query_set_1 + " (" + key_1 + ") AND (" + key_2 + ") AND (" + str(key_3) + ") " + query_set_2
#     #返信
#     response = requests.get(query)
#     #IDの解析
#     id_set = response.json()['esearchresult']['idlist']
#     #論文数
#     cnt = response.json()['esearchresult']['count']
#     #検索時間
#     tm_en = time.time()
#     total = tm_en - tm_st
#     #書き込み
#     #論文数
#     ws["H1"].value = cnt
#     #検索時間
#     ws["H2"].value = total
#     #ID
#     ws["H4"].value = id_set[0]
#     ws["H6"].value = id_set[1]
#     ws["H8"].value = id_set[2]
#     ws["H10"].value = id_set[3]
#     ws["H12"].value = id_set[4]
#     #保存する
#     # wb.save("C://Users//makyua//anaconda3//envs//gpu//Scripts//test//pubmed.xlsm")


#
#https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db={database}&term={query}
#database: 検索するデータベースを指定する
#query: 検索するキーワードを指定
#retmax: 最大の結果数
#retmode: 結果の形式
#リファレンス: https://www.ncbi.nlm.nih.gov/books/NBK25499/#chapter4.ESearch
#pt: 出版形式
import xlwings as xw
import requests
import json
import time
import openpyxl
from lxml import etree
import math

#pubmedから検索する部分
def detect(ws):
    query_set_1 = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=pubmed&term="
    query_set_2 = "&retmax=5&retmode=json&retstart="
    # query_set_2 = "AND review[pt]&retmax=5&retmode=json"
    #キーワードの取得
    key_1 = ws["B2"].value
    key_2 = ws["C2"].value
    key_3 = ws["D2"].value
    #ページ数の取得→表示させたいPMIDの番数
    page = 5*(math.floor(ws["B4"].value) - 1)
    #検索を始める
    query = query_set_1 + " (" + str(key_1) + ") AND (" + str(key_2) + ") AND (" + str(key_3) + ") " + query_set_2 + str(page)
    #検索結果の返信
    response = requests.get(query)
    #検索結果を返す
    return response
    
#PubmedIDの取得・書き込み
def pm_id(response,ws):
    #IDの取得
    id_set = response.json()['esearchresult']['idlist']
    #ID
    ws["H4"].value = id_set[0]
    ws["H6"].value = id_set[1]
    ws["H8"].value = id_set[2]
    ws["H10"].value = id_set[3]
    ws["H12"].value = id_set[4]

#論文数の取得・書き込み
def pm_cnt(response,ws):
    #論文数
    cnt = response.json()['esearchresult']['count']
    #論文数
    ws["H1"].value = cnt
        
#abstractの取得・書き込み    
def abst(response,ws):
    #IDの取得
    id_set = response.json()['esearchresult']['idlist']
    
    #abstractを入れる入れ物（変数）
    abst_list = []
    
    #クエリの作成
    que = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?db=pubmed&retmode=XML&id='
    for id_ in id_set:
        query = que + id_
        #検索
        res_xml = requests.get(query)
        #パース
        root = etree.fromstring(res_xml.content)
        #abstract部分の解析、抽出
        abst_list.append(''.join(root.xpath('//Abstract//*/text()')))
        
    #EXCELへの記入
    i = 5
    for ab in abst_list:
        ws["H"+str(i)].value = ab
        i += 2
    
    
def main():        
    #リファレンス: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.reader.excel.html#openpyxl.reader.excel.load_workbook
    EXCEL_PATH = "C:\\Users\\makyua\\Desktop\\article-3\\article\\windows\\pubmed_3.xlsm"
    #Excelファイルの読み込み
    wb = xw.Book(EXCEL_PATH)
    #Excelシートの読み込み
    ws = wb.sheets["Sheet1"]
    
    tm_st = time.time()
    
    #検索の実行、検索結果の格納
    response = detect(ws)
    
    #PMIDの取得、書きこみ
    pm_id(response,ws)
    
    #論文数の取得、書き込み
    pm_cnt(response,ws)
    
    #abstractの取得、書き込み
    abst(response,ws)
    
    #検索時間の算出
    tm_en = time.time()
    total = tm_en - tm_st
    #検索時間の記述
    ws["H2"].value = total
    #保存する
    # wb.save("C://Users//makyua//anaconda3//envs//gpu//Scripts//test//pubmed.xlsm")