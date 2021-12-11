import xlwings as xw
import requests
# import json
import time
import openpyxl
from lxml import etree
from itertools import product

#pubmedから検索する部分
def detect(kensaku):
    query_set_1 = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=pubmed&term="
    query_set_2 = "&retmode=json"
    #組み合わせの取得: 6*5*5=150通り
    keylist = [list(v) for v in product(*kensaku)]
    #組み合わせからクエリを取得
    query = [query_set_1 + " (" + str(q[0]) + ") AND (" + str(q[1]) + ") AND (" + str(q[2]) + ") " + query_set_2 for q in keylist]
    #検索結果の返信
    response = []
    for res in query:
        response.append(requests.get(res))
        print(res)
#     print(query[0])
    #検索結果を返す
    return response, keylist

def keyword(ws):
    #キーワード数
    cnt = 3
    #キーワードボックス
    kensaku = [[], [], []]
    #キーワードの取得
    num = 2
    while ws["b"][num].value:
        kensaku[0].append(ws["b"][num].value)
        num+=1
    num = 2
    while ws["c"][num].value:
        kensaku[1].append(ws["c"][num].value)
        num+=1
    num = 2
    while ws["d"][num].value:
        kensaku[2].append(ws["d"][num].value)
        num+=1
    return kensaku

#論文数の取得・書き込み
def pm_cnt(response):
    #論文数
    cnt = [res.json()['esearchresult']['count'] for res in response]
    return cnt

def show(cnt, keylist, ws):
    i = 1
    for (c, k) in zip(cnt, keylist):
        ws.cell(row=15+i, column=7).value=i
        ws.cell(row=15+i, column=8).value=k[0]
        ws.cell(row=15+i, column=9).value=k[1]
        ws.cell(row=15+i, column=10).value=k[2]
        ws.cell(row=15+i, column=11).value=int(c)
        i+=1
        

#リファレンス: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.reader.excel.html#openpyxl.reader.excel.load_workbook
EXCEL_PATH = "C:\\Users\\makyua\\Desktop\\article-3\\article\\windows\\pubmed_3.xlsm"
#Excelファイルの読み込み
wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
#Excelシートの読み込み
ws = wb["Season2"]

tm_st = time.time()

#検索キーワードの取得
kensaku = keyword(ws)

# #検索の実行、検索結果の格納
response, keylist = detect(kensaku)

# #検索結果から論文数の解析
cnt = pm_cnt(response)

#データを結合後、エクセルに反映
show(cnt, keylist, ws)

wb.save(EXCEL_PATH)